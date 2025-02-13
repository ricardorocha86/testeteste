import streamlit as st
from openai import OpenAI
from pydantic import BaseModel
from openpyxl import load_workbook
from typing import Literal
import pandas as pd
import os

def buscar_aulas(diretorio):
    return [f for f in os.listdir(diretorio) if f.endswith('.txt')]

def ler_aula(arquivo):
    with open(arquivo, 'r', encoding='utf-8') as f:
        return f.read()

def aux(x):
    return x[:-4]


##########
def gerar_listinha(conteudo):

    instrucoes = """
    Você é um assistente especializado em análise de texto educacional, responsável por identificar os principais temas abordados 
    em um conteúdo fornecido. Para cada texto recebido, extraia exatamente cinco temas centrais que representem de forma clara e objetiva 
    os principais tópicos discutidos. Esses temas devem ser concisos, diretos e fiéis ao conteúdo, servindo como base para a formulação de 
    perguntas relacionadas ao material. A partir desses temas, gere uma lista de exatamente cinco perguntas que explorem o conteúdo de maneira 
    relevante e coerente. Cada vez que esse processo for executado, a resposta deve ser única, variando a forma de abordagem e formulação das 
    perguntas, sem repetir estruturas fixas. O output deve conter apenas essa lista de cinco perguntas, sem qualquer outro elemento adicional.   
    """


    class Perguntas(BaseModel):
        perguntas: list[str]


    completion = client.beta.chat.completions.parse(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": instrucoes},
            {"role": "user", "content": conteudo},
        ],
        response_format=Perguntas,
    )

    event = completion.choices[0].message.parsed
    listinha = dict(event)['perguntas']
    return listinha

def gerar_perguntas(tema):
    class Questao(BaseModel):
        pergunta: str
        resposta1: str
        resposta2: str
        resposta3: str
        resposta4: str
        tempo: Literal[20]
        correta: Literal[1, 2, 3, 4]


    instrucoes = """
    Sempre que receber um tema, sua tarefa será criar uma pergunta no estilo quiz, formulada de maneira clara, objetiva e diretamente
    relacionada ao tema fornecido. A pergunta deve ter, no máximo, 120 caracteres. Além disso, gere quatro alternativas de resposta, 
    sendo uma correta e três falsas. Todas as opções devem ser coerentes com o contexto educacional e desafiadoras o suficiente para 
    estimular o pensamento crítico. Cada resposta deve ter, no máximo, 70 caracteres. O tempo padrão para responder a cada pergunta 
    será de 20 segundos.    
    """

    completion = client.beta.chat.completions.parse(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": instrucoes},
            {"role": "user", "content": tema},
        ],
        response_format=Questao,
    )

    event = completion.choices[0].message.parsed

    return dict(event)

def GeraKahoot(conteudo):

    listinha = gerar_listinha(conteudo)


    dados = []
    for tema in listinha:
        resultado = gerar_perguntas(tema)
        dados.append(resultado)

    df = pd.DataFrame(dados)
    df.to_excel('perguntas.xlsx', index=False)

    return df


##########

# Configuração principal
diretorio_aulas = 'Aulas'
aulas = buscar_aulas(diretorio_aulas)

# Widgets
aula_selecionada = st.sidebar.selectbox(
    'Selecione uma aula:', 
    sorted(aulas), 
    index=0, 
    format_func=aux
)


 
def GeraExcel(kahoot):

    excel_path = 'Kahoot-Quiz-Spreadsheet-Template.xlsx'

    # Carregar o arquivo Excel com openpyxl
    wb = load_workbook(excel_path)
    st.write('carregou')
    
    # Escolher a planilha ativa (ou usar wb['nome_da_planilha'])
    ws = wb.active
    st.write('carregou2')

    # Definir a linha e a coluna onde começar a escrever
    linha_inicial = 9
    coluna_inicial = 2  # Isso será a coluna "C" (1 = A, 2 = B, 3 = C, etc.)

    # Escrever o DataFrame no arquivo Excel a partir da célula definida
    for i, row in kahoot.iterrows():
        for j, value in enumerate(row):
            # A linha e a coluna começam a partir dos índices definidos
            ws.cell(row=linha_inicial + i, column=coluna_inicial + j, value=value)
    st.write('carregou3')

    # Salvar o arquivo Excel modificado
    wb.save('ExcelKahoot.xlsx')
    st.write('carregou4')



# Carregar conteúdo
caminho_aula = os.path.join(diretorio_aulas, aula_selecionada)
conteudo_aula = ler_aula(caminho_aula)
openai_api_key = st.secrets["OPENAI_API_KEY"]  # Garantir que está usando o segredo
client = OpenAI(api_key=openai_api_key)  
# Botão de geração
st.write('Oie - aqui é um gerador de Kahoot - issso faz assim assim assado ')
if st.button('✨ Gerar Kahoot', type='primary'):
    kahoot = GeraKahoot(conteudo_aula)
    st.write(kahoot)
    #GeraExcel(kahoot)
    # Crie o botão de download
    # Abra o arquivo Excel em modo binário
    #with open('ExcelKahoot.xlsx', 'rb') as file:
     #   excel_data = file.read()
 
    kahoot.rename(columns={
    'pergunta': 'Question - max 120 characters',
    'resposta1': 'Answer 1 - max 75 characters',
    'resposta2': 'Answer 2 - max 75 characters',
    'resposta3': 'Answer 3 - max 75 characters',
    'resposta4': 'Answer 4 - max 75 characters',
    'tempo': 'Time limit (sec) – 5, 10, 20, 30, 60, 90, 120, or 240 secs',
    'correta': 'Correct answer(s) - choose at least one'
}, inplace=True)

 
    def to_excel(df):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
        processed_data = output.getvalue()
        return processed_data

    # Convertendo o DataFrame para Excel
    excel_data = to_excel(kahoot)

    # Criando o botão de download
    st.download_button(
        label="Baixar Kahoot",
        data=excel_data,
        file_name="kahoot.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
